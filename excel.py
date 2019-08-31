import openpyxl
import time,os
from openpyxl.styles import Font, colors, Alignment  #这里有个规则，不能是平级
os.chdir('D:\\learnPY\\excel')
path='excel_demo.xlsx'


#打开excel文件
wb=openpyxl.load_workbook(path)

ss=wb.get_sheet_names()
print('获取工作簿中的所有所有工作表名：',ss)

s1=wb.get_sheet_names()[0]
print('方法1：获取某个工作表名',s1)

s2=wb.get_sheet_by_name('Sheet1')
print('方法2：获取某个工作表名的对象，供后续调用其它方法使用',s2)


#获取单元格数据
print('方法1：获取单元格A1的值',s2['A1'].value)

#这个写法的好处是方便用for循环取值
print('方法2：获取单元格A2的值',s2.cell(row=2,column=1).value)

#写入数据到excel
s2.cell(row=5,column=1).value='我是一个好人'
s2['B5'].value='你呢'  #此处也可以不用value


print('获取最大列数：',s2.max_column)
print('获取最大行数：',s2.max_row)


#循环输出单元格里的数据
for i in range(1,s2.max_row+1):
    for j in range(1,s2.max_column+1):
        print(s2.cell(row=i,column=j).value,'\t',end='')  #\t，是Tab，end=''表示不换行
    print() #表示换行


#设置字体颜色
myfont=Font(color=colors.RED)
s2['B2'].font=myfont



#写入数据到excel
s2.cell(row=6,column=1).value='黄继光'
s2['B6'].value='哪咤'  #此处也可以不用value


#如果写入了数据，记得一定要保存
wb.save(path)

