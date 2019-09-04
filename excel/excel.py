import openpyxl
from openpyxl.styles import Font, colors, Alignment

print('--------------------------------------openpyxl操作excel-------------------------------------')

# 打开excel文件
wb = openpyxl.load_workbook('excel_test.xlsx')  # 重要1

# 获取所有工作表的列表
sheets = wb.sheetnames
print(sheets)

# 获取某个工作表的对象，供后续调用其它方法使用  # 重要2
sheet2 = wb['Sheet2']

# 获取最大行数、列数
print('获取最大列数：', sheet2.max_column)
print('获取最大行数：', sheet2.max_row)

# 设置字体颜色
font = Font(color=colors.RED)
alignment = Alignment(horizontal='general', vertical='center')

sheet2['B2'].font = font
sheet2['B2'].alignment = alignment

print('----------------------------写入数据到excel-------------------------------------------')
# 写入单条
sheet2['A9'].value = '美元'                           # .value可以省略
sheet2.cell(row=8, column=1).value = '我是一个好人'   # 推荐


# for循环写入
def aaa(name, age):
    for i in range(10, 15):
        sheet2.cell(row=i, column=1).value = name[i-10]
        sheet2.cell(row=i, column=2).value = age[i - 10]


name = ['jerry', 'tom', 'lucy', 'lily', 'howard']
age = [18, 19, 20, 21, 22]
aaa(name, age)

print('-------------------------获取单元格数据------------------------------------------------')
# 获取某个单元格
print(sheet2['A1'].value)
print(sheet2.cell(row=1, column=2).value)

# for循环输出所有单元格
for i in range(1, sheet2.max_row+1):
    for j in range(1, sheet2.max_column+1):
        print(sheet2.cell(row=i, column=j).value, end='\t')
    print()

print('---------------------遍历excel(模拟遍历测试用例)---------------------------------------------')
for i in range(2, sheet2.max_row+1):
    book = sheet2.cell(row=i, column=1).value
    major = sheet2.cell(row=i, column=2).value
    student = sheet2.cell(row=i, column=3).value
    score = sheet2.cell(row=i, column=4).value

    print(book, major, student, score)

# 记得一定要保存
wb.save('excel_test.xlsx')
