import openpyxl


def clear_result(s):
    for i in range(2, s.max_row + 1):
        s.cell(row=i, column=11).value = ''
        s.cell(row=i, column=12).value = ''
        s.cell(row=i, column=13).value = ''

    print('清空excel表中测试结果、响应数据、响应时间!')


def main(test_case):
    wb = openpyxl.load_workbook(test_case)
    s = wb['TestCase']
    clear_result(s)

    wb.save(test_case)
    return wb, s
