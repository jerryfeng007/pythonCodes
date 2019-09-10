from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import Alignment

# 设置Excel中单元格的样式
myf = Font(color=colors.GREEN)
myf2 = Font(color=colors.RED)
mya = Alignment(horizontal='center', vertical='center')


# 用例成功时，把响应数据插入excel
def case_pass(s,
              num,
              response,
              time,
              name,
              status_code,
              log):

    s.cell(row=num + 1, column=11).value = '成功'
    s.cell(row=num + 1, column=11).font = myf
    s.cell(row=num + 1, column=11).alignment = mya
    s.cell(row=num + 1, column=13).value = time
    s.cell(row=num + 1, column=12).value = str(response)  # 需要把字典类型转化为字符类型

    log.info('编号=' + str(num) + ' ' +
             '接口名称=' + name + ',成功' + ' ' +
             'http状态码=' + str(status_code) + ' ' +
             '响应时间=' + str(time) + '秒' + '\n' +
             '响应内容=' + str(response))


# 用例失败时，把响应数据插入excel
def case_fail(s,
              num,
              response,
              time,
              name,
              status_code,
              log):

    s.cell(row=num + 1, column=11).value = '失败'
    s.cell(row=num + 1, column=11).font = myf2
    s.cell(row=num + 1, column=11).alignment = mya
    s.cell(row=num + 1, column=13).value = time
    s.cell(row=num + 1, column=12).value = str(response)

    log.info('编号=' + str(num) + ' ' +
             '接口名称=' + name + ',失败' + ' ' +
             'http状态码=' + str(status_code) + ' ' +
             '响应时间=' + str(time) + '秒' + '\n' +
             '响应内容=' + str(response))
