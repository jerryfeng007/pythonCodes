import time
import requests
import openpyxl
from bs4 import BeautifulSoup


def deal(target_url, headers, file, base_url='http://www.xqtesting.com'):
    res = requests.get(target_url, headers=headers).content

    # 创建对象，以python内置的标准库
    soup = BeautifulSoup(res, 'html.parser')

    # find_all方法搜索所有满足条件的tag
    h4 = soup.find_all('h4', {'class': 'card-heading'})
    # print(h4)

    for link in h4:
        title = link.a.get_text()
        print(title)
        file.write(title+'\n')
        full_url = base_url+link.a['href']
        print(full_url.strip())  # strip()去掉空格
        file.write(full_url.strip()+'\n')


def urls():
    urls = []
    for i in range(1, table.max_row+1):
        target_url = table.cell(row=i, column=1).value.replace('\n', '').replace('\r', '')
        urls.append(target_url)

    with open('D:\\learnPY\\crawler\\blog.txt', 'w') as file:
        for url in urls:
            deal(url, headers, file)


if __name__ == '__main__':
    s = time.time()

    # 设置headers能够更好的模拟浏览器访问，防止被禁用
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64; rv:24.0) Gecko/20100101 Firefox/24.0'}

    wb = openpyxl.load_workbook('D:\\learnPY\\crawler\\target_url.xlsx')
    table = wb['Sheet1']

    urls()

    e = time.time()
    print(e - s)
