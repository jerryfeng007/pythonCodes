import time
import openpyxl
from bs4 import BeautifulSoup
import aiohttp
import asyncio


async def deal(target_url, headers, file, base_url='http://www.xqtesting.com'):
    async with aiohttp.ClientSession() as s:
        async with s.get(target_url, headers=headers) as res:
            res = await res.text()

    # 创建对象，以python内置的标准库
    soup = BeautifulSoup(res, 'html.parser')

    # find_all方法搜索所有满足条件的tag
    h4 = soup.find_all('h4', {'class': 'card-heading'})
    # print(h4)

    for link in h4:
        title = link.a.get_text()
        print(title)
        file.write(title+'\n')
        full_url = base_url+link.a['href']
        print(full_url.strip())  # strip()去掉空格
        file.write(full_url.strip()+'\n')


async def urls():
    urls = []
    for i in range(1, table.max_row+1):
        target_url = table.cell(row=i, column=1).value.replace('\n', '').replace('\r', '')
        urls.append(target_url)

    with open('blog_异步.txt', 'w') as file:
        tasks = [asyncio.create_task(deal(url, headers, file)) for url in urls]
        await asyncio.gather(*tasks)


if __name__ == '__main__':
    s = time.time()

    # 设置headers能够更好的模拟浏览器访问，防止被禁用
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64; rv:24.0) Gecko/20100101 Firefox/24.0'}

    wb = openpyxl.load_workbook('target_url.xlsx')
    table = wb['Sheet1']

    asyncio.run(urls())

    e = time.time()
    print(e - s)
